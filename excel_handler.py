# -*- coding: utf-8 -*-
"""
Excel数据处理模块
提供AHP和模糊综合评价所需的Excel数据读取与导出功能
支持专家判断矩阵、评分数据、权重信息的处理
"""

import os
import time
import pandas as pd
import numpy as np
import logging
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple, Optional, Any, Union
from fractions import Fraction
import matplotlib.pyplot as plt
import seaborn as sns


class ExcelExporter:
    """
    Excel文件导出器，用于格式化输出分析结果
    支持AHP结果、模糊评价结果和敏感性分析结果的导出
    """

    def __init__(self):
        """初始化Excel导出器"""
        self.logger = logging.getLogger(__name__)

        # 预定义样式
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.modified_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        self.header_font = Font(color="FFFFFF", bold=True)
        self.subheader_font = Font(bold=True)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def adjust_all_sheets(self, writer: pd.ExcelWriter) -> None:
        """
        调整所有工作表的列宽和对齐方式

        Args:
            writer (pd.ExcelWriter): Excel写入器
        """
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            # 调整列宽
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)

                # 设置列宽（最大内容长度 + 4个字符的边距）
                adjusted_width = max(max_length + 4, 10)  # 最小宽度为10
                worksheet.column_dimensions[column_letter].width = adjusted_width

                # 设置单元格对齐方式
                for cell in column_cells:
                    cell.alignment = self.center_alignment

    def matrix_compare_and_highlight(self,
                                     original_matrix: np.ndarray,
                                     corrected_matrix: np.ndarray,
                                     writer: pd.ExcelWriter,
                                     sheet_name: str,
                                     criteria: List[str]) -> None:
        """
        比较并高亮显示原始和修正矩阵的差异

        Args:
            original_matrix (np.ndarray): 原始判断矩阵
            corrected_matrix (np.ndarray): 修正后的判断矩阵
            writer (pd.ExcelWriter): Excel写入器
            sheet_name (str): 工作表名称
            criteria (List[str]): 准则名称列表
        """
        n = len(original_matrix)

        # 创建原始矩阵DataFrame
        df1 = pd.DataFrame(np.round(original_matrix, 4), columns=criteria, index=criteria)
        df1.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=0)

        # 确定修正矩阵的起始列
        offset = n + 3

        # 创建修正矩阵DataFrame
        df2 = pd.DataFrame(np.round(corrected_matrix, 4), columns=criteria, index=criteria)
        df2.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=offset)

        # 获取工作表
        ws = writer.sheets[sheet_name]

        # 添加标题
        ws["A1"] = "原始判断矩阵"
        ws["A1"].font = self.subheader_font
        ws["A1"].fill = self.subheader_fill

        # 获取修正矩阵标题单元格位置
        title_cell = f"{get_column_letter(offset + 1)}1"
        ws[title_cell] = "修正矩阵"
        ws[title_cell].font = self.subheader_font
        ws[title_cell].fill = self.subheader_fill

        # 高亮显示修正的单元格
        for i in range(n):
            for j in range(n):
                if abs(original_matrix[i, j] - corrected_matrix[i, j]) > 1e-4:
                    # 原始矩阵中的单元格
                    cell1 = f"{get_column_letter(j + 2)}{i + 3}"
                    ws[cell1].fill = self.modified_fill

                    # 修正矩阵中的单元格
                    cell2 = f"{get_column_letter(offset + j + 2)}{i + 3}"
                    ws[cell2].fill = self.modified_fill

        # 添加一致性比率信息
        from ahp_processor import ConsistencyChecker

        checker = ConsistencyChecker()
        original_cr = checker.check_consistency(original_matrix).consistency_ratio
        corrected_cr = checker.check_consistency(corrected_matrix).consistency_ratio

        # 在原始矩阵下方添加CR信息
        cr_row = n + 4
        ws[f"A{cr_row}"] = "一致性比率 (CR)"
        ws[f"B{cr_row}"] = round(original_cr, 4)

        # 在修正矩阵下方添加CR信息
        ws[f"{get_column_letter(offset + 1)}{cr_row}"] = "一致性比率 (CR)"
        ws[f"{get_column_letter(offset + 2)}{cr_row}"] = round(corrected_cr, 4)

        # 添加一致性状态信息
        status_row = cr_row + 1
        ws[f"A{status_row}"] = "一致性状态"
        ws[f"B{status_row}"] = "一致" if original_cr <= 0.1 else "不一致"

        ws[f"{get_column_letter(offset + 1)}{status_row}"] = "一致性状态"
        ws[f"{get_column_letter(offset + 2)}{status_row}"] = "一致" if corrected_cr <= 0.1 else "不一致"

    def export_ahp_results(self,
                           results: Dict[str, Any],
                           output_path: str,
                           include_matrices: bool = True) -> None:
        """
        导出AHP分析结果到Excel文件

        Args:
            results (Dict[str, Any]): AHP分析结果
            output_path (str): 输出文件路径
            include_matrices (bool): 是否包含详细的判断矩阵
        """
        # 添加时间戳，避免覆盖文件
        timestamp = int(time.time())
        output_path = f"{os.path.splitext(output_path)[0]}_{timestamp}.xlsx"

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. 汇总信息
                summary_data = {
                    "分析指标": [
                        "专家数量",
                        "评价准则数量",
                        "群组矩阵一致性比率(CR)",
                        "修正判断矩阵数量",
                        "分析日期",
                        "分析时间"
                    ],
                    "结果值": [
                        len(results["corrected_matrices"]),
                        len(results["final_weights"]),
                        round(results["aggregated_consistency"].consistency_ratio, 4),
                        sum(1 for corr in results["correction_results"] if corr.adjusted),
                        time.strftime("%Y-%m-%d"),
                        time.strftime("%H:%M:%S")
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name="汇总信息", index=False)

                # 2. 最终权重
                weights_data = [
                    {"准则名": criterion, "权重值": round(weight, 4)}
                    for criterion, weight in sorted(
                        results["final_weights"].items(),
                        key=lambda x: x[1],
                        reverse=True
                    )
                ]
                pd.DataFrame(weights_data).to_excel(writer, sheet_name="最终权重", index=False)

                # 3. 一致性分析结果
                consistency_data = [
                    {
                        "专家编号": f"专家{corr.expert_id}",
                        "原始CR值": round(corr.original_cr, 4),
                        "修正后CR值": round(corr.final_cr, 4),
                        "是否达到一致性": "是" if corr.success else "否",
                        "是否动态修正": "是" if corr.adjusted else "否",
                        "修正方法": corr.correction_method if hasattr(corr, 'correction_method') else "LLSM"
                    }
                    for corr in results["correction_results"]
                ]
                pd.DataFrame(consistency_data).to_excel(writer, sheet_name="专家一致性分析结果", index=False)

                # 4. 群组判断矩阵
                if include_matrices:
                    criteria = list(results["final_weights"].keys())
                    aggregated_df = pd.DataFrame(
                        np.round(results["aggregated_matrix"], 4),
                        index=criteria,
                        columns=criteria
                    )
                    aggregated_df.to_excel(writer, sheet_name="群组判断矩阵")

                    # 5. 专家判断矩阵比较
                    for i, (original, corrected) in enumerate(
                            zip(results["original_matrices"], results["corrected_matrices"])
                    ):
                        sheet_name = f"Expert{i + 1}_Comparison"
                        self.matrix_compare_and_highlight(
                            original, corrected, writer, sheet_name, criteria
                        )

                # 6. 添加权重可视化
                self._add_weights_visualization(writer, results["final_weights"])

                # 调整所有工作表的格式
                self.adjust_all_sheets(writer)

            self.logger.info(f"已导出AHP分析结果到: {output_path}")

        except Exception as e:
            self.logger.error(f"导出AHP结果时出错: {str(e)}")
            raise

    def export_fuzzy_results(self,
                             fuzzy_results: Dict[str, Any],
                             output_path: str) -> None:
        """
        导出模糊综合评价结果到Excel文件

        Args:
            fuzzy_results (Dict[str, Any]): 模糊评价结果
            output_path (str): 输出文件路径
        """
        # 添加时间戳，避免覆盖文件
        timestamp = int(time.time())
        output_path = f"{os.path.splitext(output_path)[0]}_{timestamp}.xlsx"

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. 模糊评价结果（隶属度）
                if "integrated_result" in fuzzy_results:
                    result = fuzzy_results["integrated_result"]
                    fuzzy_df = pd.DataFrame(
                        [result.round(4)],
                        index=["风险等级"],
                        columns=["VL", "L", "M", "H", "VH"]
                    )
                    fuzzy_df.to_excel(writer, sheet_name="模糊评价结果")

                # 2. 风险指数
                if "risk_index" in fuzzy_results:
                    risk_index = fuzzy_results["risk_index"]

                    # 风险等级判定
                    if risk_index < 0.3:
                        risk_level = "低风险"
                    elif risk_index < 0.5:
                        risk_level = "中低风险"
                    elif risk_index < 0.7:
                        risk_level = "中风险"
                    elif risk_index < 0.9:
                        risk_level = "中高风险"
                    else:
                        risk_level = "高风险"

                    index_data = {
                        "指标": ["风险指数", "风险等级"],
                        "值": [round(risk_index, 4), risk_level]
                    }
                    pd.DataFrame(index_data).to_excel(writer, sheet_name="风险指数", index=False)

                # 3. 因素隶属度
                if "factor_membership" in fuzzy_results:
                    factor_membership = fuzzy_results["factor_membership"]
                    membership_data = []

                    for factor, membership in factor_membership.items():
                        row_data = {"风险因素": factor}
                        for i, level in enumerate(["VL", "L", "M", "H", "VH"]):
                            row_data[level] = round(membership[i], 4)
                        membership_data.append(row_data)

                    membership_df = pd.DataFrame(membership_data)
                    membership_df.to_excel(writer, sheet_name="因素隶属度", index=False)

                # 4. 敏感性分析结果
                if "sensitivity_analysis" in fuzzy_results:
                    sensitivity = fuzzy_results["sensitivity_analysis"]

                    # 敏感性指标
                    sens_data = []
                    for factor, value in sensitivity["sensitivity_indices"].items():
                        sens_data.append({"风险因素": factor, "敏感性指标": round(value, 4)})

                    sens_df = pd.DataFrame(sens_data)
                    sens_df = sens_df.sort_values(by="敏感性指标", ascending=False)
                    sens_df.to_excel(writer, sheet_name="敏感性指标", index=False)

                    # 关键风险因素
                    if "critical_factors" in sensitivity:
                        critical_df = pd.DataFrame({"关键风险因素": sensitivity["critical_factors"]})
                        critical_df.to_excel(writer, sheet_name="关键风险因素", index=False)

                # 5. 添加可视化
                self._add_fuzzy_visualization(writer, fuzzy_results)

                # 调整所有工作表的格式
                self.adjust_all_sheets(writer)

            self.logger.info(f"已导出模糊评价结果到: {output_path}")

        except Exception as e:
            self.logger.error(f"导出模糊评价结果时出错: {str(e)}")
            raise

    def _add_weights_visualization(self, writer: pd.ExcelWriter, weights: Dict[str, float]) -> None:
        """
        添加权重可视化到Excel文件

        Args:
            writer (pd.ExcelWriter): Excel写入器
            weights (Dict[str, float]): 权重字典
        """
        # 创建一个新的工作表
        workbook = writer.book
        sheet_name = "权重可视化"

        if sheet_name in workbook.sheetnames:
            # 如果工作表已存在，删除它
            idx = workbook.sheetnames.index(sheet_name)
            workbook.remove(workbook.worksheets[idx])

        worksheet = workbook.create_sheet(sheet_name)

        # 创建权重条形图
        fig, ax = plt.figure(figsize=(10, 6)), plt.axes()

        # 对权重进行排序
        sorted_weights = sorted(weights.items(), key=lambda x: x[1], reverse=True)
        factors = [item[0] for item in sorted_weights]
        values = [item[1] for item in sorted_weights]

        # 绘制条形图
        bars = ax.barh(factors, values, color='skyblue')

        # 添加数据标签
        for bar in bars:
            width = bar.get_width()
            ax.text(width + 0.01, bar.get_y() + bar.get_height() / 2, f'{width:.4f}',
                    va='center')

        # 设置图表标题和标签
        ax.set_title('准则权重')
        ax.set_xlabel('权重值')
        ax.set_ylabel('准则')

        # 调整布局
        plt.tight_layout()

        # 保存图表到临时文件
        temp_file = 'temp_weights_chart.png'
        plt.savefig(temp_file, dpi=300)
        plt.close()

        # 将图表插入到Excel
        from openpyxl.drawing.image import Image
        img = Image(temp_file)
        img.width = 600
        img.height = 400
        worksheet.add_image(img, 'A1')

        # 删除临时文件
        try:
            os.remove(temp_file)
        except:
            pass

    def _add_fuzzy_visualization(self, writer: pd.ExcelWriter, fuzzy_results: Dict[str, Any]) -> None:
        """
        添加模糊评价可视化到Excel文件

        Args:
            writer (pd.ExcelWriter): Excel写入器
            fuzzy_results (Dict[str, Any]): 模糊评价结果
        """
        # 创建一个新的工作表
        workbook = writer.book
        sheet_name = "评价可视化"

        if sheet_name in workbook.sheetnames:
            # 如果工作表已存在，删除它
            idx = workbook.sheetnames.index(sheet_name)
            workbook.remove(workbook.worksheets[idx])

        worksheet = workbook.create_sheet(sheet_name)

        # 创建风险等级条形图
        if "integrated_result" in fuzzy_results:
            fig, ax = plt.figure(figsize=(10, 6)), plt.axes()

            result = fuzzy_results["integrated_result"]
            levels = ["VL", "L", "M", "H", "VH"]

            # 绘制条形图
            bars = ax.bar(levels, result, color=['blue', 'green', 'yellow', 'orange', 'red'])

            # 添加数据标签
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, height + 0.01, f'{height:.4f}',
                        ha='center')

            # 设置图表标题和标签
            ax.set_title('风险等级隶属度分布')
            ax.set_xlabel('风险等级')
            ax.set_ylabel('隶属度')

            # 调整布局
            plt.tight_layout()

            # 保存图表到临时文件
            temp_file = 'temp_fuzzy_chart.png'
            plt.savefig(temp_file, dpi=300)
            plt.close()

            # 将图表插入到Excel
            from openpyxl.drawing.image import Image
            img = Image(temp_file)
            img.width = 600
            img.height = 400
            worksheet.add_image(img, 'A1')

            # 删除临时文件
            try:
                os.remove(temp_file)
            except:
                pass

        # 如果有敏感性分析结果，添加敏感性图表
        if "sensitivity_analysis" in fuzzy_results:
            sensitivity = fuzzy_results["sensitivity_analysis"]

            if "sensitivity_indices" in sensitivity:
                # 创建一个新的工作表
                sheet_name = "敏感性可视化"

                if sheet_name in workbook.sheetnames:
                    # 如果工作表已存在，删除它
                    idx = workbook.sheetnames.index(sheet_name)
                    workbook.remove(workbook.worksheets[idx])

                worksheet = workbook.create_sheet(sheet_name)

                # 创建敏感性条形图
                fig, ax = plt.figure(figsize=(10, 6)), plt.axes()

                # 排序敏感性指标
                sorted_indices = sorted(
                    sensitivity["sensitivity_indices"].items(),
                    key=lambda x: abs(x[1]),
                    reverse=True
                )

                # 只显示前10个因素
                if len(sorted_indices) > 10:
                    sorted_indices = sorted_indices[:10]

                factors = [item[0] for item in sorted_indices]
                values = [item[1] for item in sorted_indices]

                # 为正负值设置不同颜色
                colors = ['blue' if v >= 0 else 'red' for v in values]

                # 绘制条形图
                bars = ax.barh(factors, values, color=colors)

                # 添加数据标签
                for bar in bars:
                    width = bar.get_width()
                    ax.text(width + 0.01 if width >= 0 else width - 0.05,
                            bar.get_y() + bar.get_height() / 2,
                            f'{width:.4f}',
                            va='center')

                # 设置图表标题和标签
                ax.set_title('风险因素敏感性指标')
                ax.set_xlabel('敏感性指标')
                ax.set_ylabel('风险因素')

                # 添加参考线
                ax.axvline(x=0, color='black', linestyle='-', alpha=0.3)

                # 调整布局
                plt.tight_layout()

                # 保存图表到临时文件
                temp_file = 'temp_sensitivity_chart.png'
                plt.savefig(temp_file, dpi=300)
                plt.close()

                # 将图表插入到Excel
                from openpyxl.drawing.image import Image
                img = Image(temp_file)
                img.width = 600
                img.height = 400
                worksheet.add_image(img, 'A1')

                # 删除临时文件
                try:
                    os.remove(temp_file)
                except:
                    pass


class ExcelDataHandler:
    """Excel文件数据处理器，负责读取AHP和模糊评价数据"""

    def __init__(self, matrix_validator=None):
        """
        初始化Excel数据处理器

        Args:
            matrix_validator: 矩阵验证器，用于验证读取的判断矩阵
        """
        self.matrix_validator = matrix_validator
        self.logger = logging.getLogger(__name__)

    def read_expert_matrices(self,
                             excel_path: str,
                             sheet_prefix: str = "Expert") -> Tuple[List[np.ndarray], List[str]]:
        """
        读取专家判断矩阵数据

        Args:
            excel_path (str): Excel文件路径
            sheet_prefix (str): 专家工作表前缀

        Returns:
            Tuple[List[np.ndarray], List[str]]: 专家判断矩阵列表和准则名称列表

        Raises:
            FileNotFoundError: 如果文件不存在
            ValueError: 如果数据格式不正确
        """
        # 检查文件是否存在
        if not os.path.exists(excel_path):
            error_msg = f"文件未找到: {excel_path}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)

        try:
            # 读取Excel文件
            xl = pd.ExcelFile(excel_path)

            # 找出所有专家工作表
            expert_sheets = [s for s in xl.sheet_names if s.startswith(sheet_prefix)]

            if not expert_sheets:
                error_msg = f"在 {excel_path} 中未找到以 '{sheet_prefix}' 开头的工作表"
                self.logger.error(error_msg)
                raise ValueError(error_msg)

            # 初始化结果列表
            matrices = []
            criteria_names = None

            # 处理每个专家工作表
            for sheet in expert_sheets:
                # 读取工作表数据
                df = pd.read_excel(excel_path, sheet_name=sheet, index_col=0)

                # 记录准则名称
                if criteria_names is None:
                    criteria_names = df.index.tolist()
                elif df.index.tolist() != criteria_names:
                    error_msg = f"{sheet} 中的准则名称与其他工作表不匹配"
                    self.logger.error(error_msg)
                    raise ValueError(error_msg)

                # 创建矩阵
                n = len(df)
                partial_matrix = np.zeros((n, n))

                # 填充上三角部分
                for i in range(n):
                    for j in range(i, n):
                        value = df.iloc[i, j]

                        # 处理分数形式的输入
                        if isinstance(value, str) and '/' in value:
                            try:
                                partial_matrix[i, j] = float(Fraction(value))
                            except:
                                error_msg = f"在 {sheet} 中发现无效的分数: '{value}'"
                                self.logger.error(error_msg)
                                raise ValueError(error_msg)
                        else:
                            # 处理数值形式的输入
                            partial_matrix[i, j] = pd.to_numeric(value, errors="coerce")

                            # 检查是否转换成功
                            if np.isnan(partial_matrix[i, j]):
                                error_msg = f"在 {sheet} 的单元格 ({i + 1}, {j + 1}) 中发现无效值: {value}"
                                self.logger.error(error_msg)
                                raise ValueError(error_msg)

                # 验证上三角矩阵
                if self.matrix_validator:
                    try:
                        self.matrix_validator.validate_upper_triangular(partial_matrix, sheet)
                    except ValueError as e:
                        self.logger.error(f"矩阵验证失败: {str(e)}")
                        raise

                # 补全完整矩阵
                complete_matrix = self._complete_matrix(partial_matrix)
                matrices.append(complete_matrix)

            self.logger.info(f"成功读取 {len(matrices)} 个专家判断矩阵，准则数量: {len(criteria_names)}")
            return matrices, criteria_names

        except Exception as e:
            if not isinstance(e, (FileNotFoundError, ValueError)):
                self.logger.error(f"读取专家判断矩阵时出错: {str(e)}")
            raise

    def read_expert_scores(self, excel_path: str) -> Tuple[pd.DataFrame, Optional[List[float]]]:
        """
        读取专家评分数据

        Args:
            excel_path (str): Excel文件路径

        Returns:
            Tuple[pd.DataFrame, Optional[List[float]]]: 专家评分数据框和专家权重（可选）

        Raises:
            FileNotFoundError: 如果文件不存在
            ValueError: 如果数据格式不正确
        """
        # 检查文件是否存在
        if not os.path.exists(excel_path):
            error_msg = f"文件未找到: {excel_path}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)

        try:
            # 读取评分数据
            scores_df = pd.read_excel(
                excel_path,
                sheet_name="专家评分",
                index_col=0  # 使用第一列作为索引（风险因素）
            )

            # 验证评分范围（应在1-10之间）
            if not scores_df.apply(lambda x: x.between(1, 10)).all().all():
                error_msg = "评分超出 1-10 范围"
                self.logger.error(error_msg)
                raise ValueError(error_msg)

            # 尝试读取专家权重
            expert_weights = None
            try:
                weights_df = pd.read_excel(excel_path, sheet_name="专家权重")
                expert_weights = weights_df["权重值"].values.tolist()

                # 检查权重和是否为1
                if not np.isclose(sum(expert_weights), 1.0, rtol=1e-5):
                    self.logger.warning("专家权重和不为 1, 将使用使用默认值")
            except Exception as e:
                self.logger.warning(f"无法读取专家权重: {str(e)}，将使用默认均等权重")

            self.logger.info(f"成功读取专家评分数据，风险因素数量: {len(scores_df)}")
            return scores_df, expert_weights

        except Exception as e:
            if not isinstance(e, (FileNotFoundError, ValueError)):
                self.logger.error(f"读取专家评分数据时出错: {str(e)}")
            raise

    @staticmethod
    def _complete_matrix(partial_matrix: np.ndarray) -> np.ndarray:
        """
        补全判断矩阵（根据上三角填充下三角）

        Args:
            partial_matrix (np.ndarray): 部分填充的矩阵（上三角）

        Returns:
            np.ndarray: 完整的判断矩阵
        """
        n = len(partial_matrix)
        complete = np.ones((n, n))

        # 复制上三角元素
        for i in range(n):
            for j in range(i + 1, n):
                complete[i, j] = partial_matrix[i, j]
                # 下三角元素为上三角元素的倒数
                complete[j, i] = 1.0 / partial_matrix[i, j]

        return complete

    def read_expert_weights(self,
                            excel_path: str,
                            sheet_name: str = "expertWeights") -> Optional[List[float]]:
        """
        读取专家权重

        Args:
            excel_path (str): Excel文件路径
            sheet_name (str): 工作表名称

        Returns:
            Optional[List[float]]: 专家权重列表，如果读取失败则返回None
        """
        try:
            # 读取权重数据
            df = pd.read_excel(excel_path, sheet_name=sheet_name)

            # 提取权重列
            if "Weight" in df.columns:
                weights_col = "Weight"
            elif "weight" in df.columns:
                weights_col = "weight"
            elif "权重" in df.columns:
                weights_col = "权重"
            elif "权重值" in df.columns:
                weights_col = "权重值"
            else:
                # 如果找不到合适的列名，使用第二列
                if len(df.columns) >= 2:
                    weights_col = df.columns[1]
                else:
                    error_msg = "无法识别权重列"
                    self.logger.error(error_msg)
                    return None

            # 提取权重值
            weights = df[weights_col].values.tolist()

            # 检查权重和是否为1
            if not np.isclose(sum(weights), 1.0, rtol=1e-5):
                self.logger.warning("专家权重和不为 1")
                return None

            self.logger.info(f"成功读取专家权重: {weights}")
            return weights

        except Exception as e:
            self.logger.warning(f"读取专家权重失败: {str(e)}")
            return None

    def create_template(self,
                        output_path: str,
                        criteria: List[str],
                        num_experts: int = 5,
                        is_ahp: bool = True) -> None:
        """
        创建专家评价模板Excel文件

        Args:
            output_path (str): 输出文件路径
            criteria (List[str]): 准则列表
            num_experts (int): 专家数量
            is_ahp (bool): 是否为AHP模板（True）或FCE模板（False）
        """
        try:
            # 创建Excel写入器
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                if is_ahp:
                    # 创建AHP模板
                    self._create_ahp_template(writer, criteria, num_experts)
                else:
                    # 创建FCE模板
                    self._create_fce_template(writer, criteria, num_experts)

                # 创建说明工作表
                self._create_instruction_sheet(writer, is_ahp)

                # 创建专家权重工作表
                self._create_expert_weights_sheet(writer, num_experts)

                # 调整所有工作表的格式
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column_cells in worksheet.columns:
                        max_length = 0
                        for cell in column_cells:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = max(max_length + 4, 10)
                        worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            self.logger.info(f"成功创建{'AHP' if is_ahp else 'FCE'}模板: {output_path}")

        except Exception as e:
            self.logger.error(f"创建模板时出错: {str(e)}")
            raise

    def _create_ahp_template(self,
                             writer: pd.ExcelWriter,
                             criteria: List[str],
                             num_experts: int) -> None:
        """
        创建AHP专家判断矩阵模板

        Args:
            writer (pd.ExcelWriter): Excel写入器
            criteria (List[str]): 准则列表
            num_experts (int): 专家数量
        """
        n = len(criteria)

        # 为每个专家创建工作表
        for expert_num in range(1, num_experts + 1):
            # 创建初始矩阵
            df = pd.DataFrame(np.ones((n, n)), index=criteria, columns=criteria)

            # 将DataFrame写入Excel
            sheet_name = f"Expert{expert_num}"
            df.to_excel(writer, sheet_name=sheet_name)

            # 获取工作表
            worksheet = writer.sheets[sheet_name]

            # 设置单元格格式
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            center_alignment = Alignment(horizontal="center", vertical="center")

            # 设置所有单元格的基本格式
            for row in range(2, n + 2):  # 从第2行开始（跳过索引行）
                for col in range(2, n + 2):  # 从第2列开始（跳过索引列）
                    cell = worksheet.cell(row=row, column=col)

                    # 基本边框和对齐
                    cell.border = border
                    cell.alignment = center_alignment

                    # 对角线和下三角部分设置为灰色
                    if row >= col:
                        cell.fill = gray_fill
                        if row > col:
                            cell.value = "自动计算"
                    else:
                        # 上三角部分需要填写
                        cell.value = ""

            # 添加标题和说明
            worksheet.insert_rows(1)
            worksheet.merge_cells('A1:K1')
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f"专家{expert_num}判断矩阵 - 请填写白色单元格"
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

    def _create_fce_template(self,
                             writer: pd.ExcelWriter,
                             criteria: List[str],
                             num_experts: int) -> None:
        """
        创建FCE专家评分模板

        Args:
            writer (pd.ExcelWriter): Excel写入器
            criteria (List[str]): 风险因素列表
            num_experts (int): 专家数量
        """
        # 创建评分工作表
        worksheet = writer.book.create_sheet("专家评分")

        # 添加标题
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "风险因素评分表 (1-10分)"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 添加表头
        headers = ["风险因素"]
        for i in range(1, num_experts + 1):
            headers.append(f"专家{i}")

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 添加风险因素
        for row, factor in enumerate(criteria, 3):
            cell = worksheet.cell(row=row, column=1)
            cell.value = factor
            cell.alignment = Alignment(vertical="center")

        # 设置边框和对齐
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in range(2, len(criteria) + 3):
            for col in range(1, num_experts + 2):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                if col > 1 and row > 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # 调整列宽
        worksheet.column_dimensions['A'].width = 40
        for i in range(2, num_experts + 2):
            worksheet.column_dimensions[get_column_letter(i)].width = 12

    def _create_instruction_sheet(self,
                                  writer: pd.ExcelWriter,
                                  is_ahp: bool) -> None:
        """
        创建说明工作表

        Args:
            writer (pd.ExcelWriter): Excel写入器
            is_ahp (bool): 是否为AHP模板
        """
        worksheet = writer.book.create_sheet("使用说明", 0)

        instructions = []

        if is_ahp:
            # AHP模板说明
            instructions = [
                ["层次分析法(AHP)专家判断说明", ""],
                ["", ""],
                ["填写指南:", ""],
                ["1. 本模板用于收集专家对评价指标的重要性比较判断", ""],
                ["2. 每位专家的判断在单独的工作表中完成", ""],
                ["3. 仅需填写矩阵的上三角部分（白色单元格）", ""],
                ["4. 对角线固定为1（灰色），下三角部分将自动计算（灰色）", ""],
                ["5. 判断标准：", ""],
                ["   1分：两个指标同等重要", ""],
                ["   3分：前者比后者略重要", ""],
                ["   5分：前者比后者明显重要", ""],
                ["   7分：前者比后者强烈重要", ""],
                ["   9分：前者比后者极强重要", ""],
                ["   2,4,6,8分：相邻判断的中间值", ""],
                ["   可以使用分数形式输入，如\"1/3\"表示后者比前者略重要", ""],
                ["6. 专家权重工作表中可以设置专家权重，权重之和应等于1", ""],
                ["7. 灰色单元格已锁定，请勿修改", ""],
                ["", ""],
                ["一致性要求:", ""],
                ["1. 填写的判断矩阵应尽量保持一致性", ""],
                ["2. 如果一致性比率(CR)大于0.1，系统将自动进行修正", ""],
                ["3. 修正后的矩阵将保持您判断的整体趋势，但会调整部分值以满足一致性要求", ""]
            ]
        else:
            # FCE模板说明
            instructions = [
                ["模糊综合评价专家评分说明", ""],
                ["", ""],
                ["评分标准:", ""],
                ["1-2分：风险发生概率很低，影响程度很小", ""],
                ["3-4分：风险发生概率较低，影响程度较小", ""],
                ["5-6分：风险发生概率中等，影响程度中等", ""],
                ["7-8分：风险发生概率较高，影响程度较大", ""],
                ["9-10分：风险发生概率很高，影响程度很严重", ""],
                ["", ""],
                ["填写要求:", ""],
                ["1. 请根据项目实际情况，对每个风险因素进行1-10的评分", ""],
                ["2. 评分需考虑风险发生的概率和可能造成的影响", ""],
                ["3. 评分应客观公正，避免主观偏见", ""],
                ["4. 所有风险因素均需评分，不可空缺", ""],
                ["5. 评分仅限整数或小数（最多保留一位小数）", ""],
                ["", ""],
                ["注意事项:", ""],
                ["1. 本评分将结合层次分析法(AHP)得到的权重进行模糊综合评价", ""],
                ["2. 评分结果将用于风险优先级排序和敏感性分析", ""],
                ["3. 最终风险等级将分为: 低风险、中低风险、中风险、中高风险、高风险", ""]
            ]

        # 写入说明内容
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                # 设置标题格式
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # 设置小标题格式
                elif row_data[0].endswith(":") and col_idx == 1:
                    cell.font = Font(bold=True)

        # 调整列宽
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 50

    def _create_expert_weights_sheet(self,
                                     writer: pd.ExcelWriter,
                                     num_experts: int) -> None:
        """
        创建专家权重工作表

        Args:
            writer (pd.ExcelWriter): Excel写入器
            num_experts (int): 专家数量
        """
        worksheet = writer.book.create_sheet("专家权重")

        # 添加标题
        worksheet.merge_cells('A1:D1')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "专家权重设置"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 添加表头
        headers = ["专家编号", "权重值", "专业领域", "备注"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 添加专家行
        default_weight = 1.0 / num_experts
        for i in range(1, num_experts + 1):
            # 专家编号
            cell = worksheet.cell(row=i + 2, column=1)
            cell.value = f"专家{i}"
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 默认权重
            cell = worksheet.cell(row=i + 2, column=2)
            cell.value = default_weight
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '0.00'

        # 添加权重和验证行
        sum_row = num_experts + 3
        worksheet.cell(row=sum_row, column=1).value = "权重总和"
        sum_cell = worksheet.cell(row=sum_row, column=2)

        # 创建公式计算权重总和
        weight_range = f"B3:B{num_experts + 2}"
        sum_cell.value = f"=SUM({weight_range})"
        sum_cell.font = Font(bold=True)
        sum_cell.alignment = Alignment(horizontal="center", vertical="center")
        sum_cell.number_format = '0.00'

        # 设置条件格式，如果总和不为1则显示红色
        from openpyxl.formatting.rule import CellIsRule
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # 添加权重验证说明
        worksheet.cell(row=sum_row + 1, column=1).value = "注意"
        note_cell = worksheet.cell(row=sum_row + 1, column=2)
        note_cell.value = "权重总和必须等于1"

        # 设置边框
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in range(2, num_experts + 4):
            for col in range(1, 5):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border

        # 调整列宽
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 40