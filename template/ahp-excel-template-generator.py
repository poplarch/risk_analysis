# TODO 根据AHP层级结构模型生成专家判断的Excel模板
# 1. 专家判断矩阵模板（仅需填写上三角部分）
# 2. 专家权重表
# 3. 说明sheet
# 4. 专家判断矩阵的填写说明
# 5. 专家权重的填写说明

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment


def create_ahp_template(
        output_path: str, level: str, criteria: list, num_experts: int = 3
):
    """
    创建AHP专家判断的Excel模板（仅需填写上三角部分）

    Parameters:
        output_path (str): 输出文件路径
        criteria (list): 评价准则列表
        num_experts (int): 专家数量
    """
    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    # 创建专家判断矩阵模板
    for expert_num in range(1, num_experts + 1):
        n = len(criteria)
        # 创建初始矩阵
        df = pd.DataFrame(np.ones((n, n)), index=criteria, columns=criteria)

        # 将下三角部分设置为灰色，表示无需填写
        sheet_name = f"Expert{expert_num}"
        df.to_excel(writer, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        # 设置单元格格式
        gray_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

        # 设置所有单元格的基本格式
        for row in range(2, n + 2):
            for col in range(2, n + 2):
                cell = worksheet.cell(row=row, column=col)

                # 基本边框和对齐
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                cell.alignment = Alignment(horizontal="center")

                # 对角线和下三角部分设置为灰色
                if row >= col:
                    cell.fill = gray_fill
                    if row > col:
                        cell.value = "自动计算"
                else:
                    cell.value = ""
                    # 设置为文本格式
                    cell.number_format = "@"

        # 调整列宽
        for col in worksheet.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[col[0].column_letter].width = max_length + 8

    # 创建专家权重表
    if level == "Goal":
        weights_df = pd.DataFrame(
            {
                "Expert": [f"Expert{i}" for i in range(1, num_experts + 1)],
                "Weight": [1 / num_experts] * num_experts,
            }
        )
        weights_df.to_excel(
            excel_writer=writer, sheet_name="expertWeights", index=False
        )

    # 添加说明sheet
    instructions_df = pd.DataFrame(
        {
            "填写说明": [
                "1. 本模板用于收集专家对评价指标的重要性判断",
                "2. 每位专家的判断在单独的工作表中完成",
                "3. 仅需填写矩阵的上三角部分（白色单元格）",
                "4. 对角线固定为1（灰色），下三角部分将自动计算（灰色）",
                "5. 判断标准：",
                "   1分：两个指标同等重要",
                "   3分：前者比后者略重要",
                "   5分：前者比后者明显重要",
                "   7分：前者比后者强烈重要",
                "   9分：前者比后者极强重要",
                "   2,4,6,8分：相邻判断的中间值",
                '   可以使用分数形式输入，如"1/2"表示后者比前者略重要',
                "6. ExpertWeights工作表中可以设置专家权重，权重之和应等于1",
                "7. 灰色单元格已锁定，请勿修改",
            ]
        }
    )
    instructions_df.to_excel(
        excel_writer=writer, sheet_name="Instructions", index=False
    )

    writer.close()


def main():
    # AHP 层级结构模型
    ahp_model = {
        "Goal": {
            "criteria": {
                "C1": "技术风险",
                "C2": "业务风险",
                "C3": "管理风险",
                "C4": "运维风险",
                "C5": "供应链风险",
                "C6": "合规风险",
            },
        },
        "技术风险C1": {
            "criteria": {
                "C1.1": "可靠性风险",
                "C1.2": "性能风险",
                "C1.3": "功能不完善",
                "C1.4": "安全风险",
                "C1.5": "扩展性风险",
                "C1.6": "系统集成风险",
                "C1.7": "代码质量风险",
                "C1.8": "测试质量风险",
            },
        },
        "业务风险C2": {
            "criteria": {
                "C2.1": "业务连续性风险",
                "C2.2": "业务功能缺失",
                "C2.3": "业务异常风险",
                "C2.4": "数据不完整",
            },
        },
        "管理风险C3": {
            "criteria": {
                "C3.1": "沟通风险",
                "C3.2": "人力风险",
                "C3.3": "需求风险",
                "C3.4": "进度风险",
            },
        },
        "运维风险C4": {
            "criteria": {
                "C4.1": "设备故障风险",
                "C4.2": "应急处置风险",
                "C4.3": "监控机制缺失",
                "C4.4": "数据备份风险",
            },
        },
        "供应链风险C5": {
            "criteria": {
                "C5.1": "成本风险",
                "C5.2": "合同风险",
                "C5.3": "产品维保风险",
            },
        },
        "合规风险C6": {
            "criteria": {
                "C6.1": "政策风险",
                "C6.2": "监管风险",
                "C6.3": "技术组件不合规",
            },
        },
    }

    for key, value in ahp_model.items():
        output_file = f"ahp_template_{key}.xlsx"
        criteria = [
            f"{cn_name}{en_key}" for en_key, cn_name in value["criteria"].items()
        ]
        create_ahp_template(output_file, key, criteria, num_experts=5)
        print(f"模板创建成功：{output_file}")


if __name__ == "__main__":
    main()
