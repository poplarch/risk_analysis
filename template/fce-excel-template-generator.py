"""
Generates Excel template for expert risk evaluation data collection
Includes structured sheets for expert scores and weights
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class ExcelTemplateGenerator:
    """Generates structured Excel templates for risk evaluation data collection"""
    
    def __init__(self):
        self.wb = Workbook()
        self.header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color="95B3D7",
            end_color="95B3D7",
            fill_type="solid"
        )
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_cell_styling(self, cell, is_header=False, is_subheader=False):
        """Applies consistent styling to cells"""
        cell.border = self.border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if is_header:
            cell.fill = self.header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        elif is_subheader:
            cell.fill = self.subheader_fill
            cell.font = Font(bold=True)

    def create_instruction_sheet(self):
        """Creates instruction sheet with evaluation guidelines"""
        ws = self.wb.active
        ws.title = "评分说明"
        
        instructions = [
            ["风险评价指引", ""],
            ["", ""],
            ["评分标准", "说明"],
            ["1-2", "风险发生概率很低，影响程度很小"],
            ["3-4", "风险发生概率较低，影响程度较小"],
            ["5-6", "风险发生概率中等，影响程度中等"],
            ["7-8", "风险发生概率较高，影响程度较大"],
            ["9-10", "风险发生概率很高，影响程度很严重"],
            ["", ""],
            ["注意事项:", ""],
            ["1. 请根据项目实际情况，对每个风险因素进行1-10的评分", ""],
            ["2. 评分需考虑风险发生的概率和可能造成的影响", ""],
            ["3. 评分应客观公正，避免主观偏见", ""],
            ["4. 如有特殊情况，请在备注中说明", ""]
        ]
        
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    self._apply_cell_styling(cell, is_header=True)
                elif row_idx == 3:
                    self._apply_cell_styling(cell, is_subheader=True)
                else:
                    self._apply_cell_styling(cell)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

    def create_expert_scores_sheet(self, risk_factors):
        """Creates expert evaluation scores sheet"""
        ws = self.wb.create_sheet("专家评分")
        
        # Headers
        ws.cell(row=1, column=1, value="风险因素")
        ws.merge_cells('A1:A2')
        self._apply_cell_styling(ws['A1'], is_header=True)
        
        num_experts = 5  # Default number of experts
        for i in range(num_experts):
            col = chr(66 + i)  # B, C, D, E, F
            ws[f'{col}1'] = f'专家{i+1}'
            ws[f'{col}2'] = '评分(1-10)'
            self._apply_cell_styling(ws[f'{col}1'], is_header=True)
            self._apply_cell_styling(ws[f'{col}2'], is_subheader=True)
            
        # Risk factors
        for idx, factor in enumerate(risk_factors, 3):
            ws[f'A{idx}'] = factor
            self._apply_cell_styling(ws[f'A{idx}'])
            
            # Add data validation for scores (1-10)
            for col in range(num_experts):
                cell = ws[f'{chr(66 + col)}{idx}']
                self._apply_cell_styling(cell)
                
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        for i in range(num_experts):
            ws.column_dimensions[chr(66 + i)].width = 15

    def create_expert_weights_sheet(self):
        """Creates expert weights configuration sheet"""
        ws = self.wb.create_sheet("专家权重")
        
        # Headers
        headers = ["专家编号", "权重值", "专业领域", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_cell_styling(cell, is_header=True)
            
        # Expert rows
        for i in range(5):  # Default 5 experts
            ws.cell(row=i+2, column=1, value=f"专家{i+1}")
            ws.cell(row=i+2, column=2, value=0.2)  # Default equal weights
            
            for col in range(1, 5):
                self._apply_cell_styling(ws.cell(row=i+2, column=col))
                
        # Adjust column widths
        column_widths = [15, 15, 30, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def generate_template(self, output_path: str):
        """
        Generates complete Excel template with all required sheets
        
        Parameters:
            output_path: Path where the template will be saved
        """
        # Define risk factors based on common categories
        risk_factors = [
            "技术架构风险",
            "数据迁移风险",
            "系统整合风险",
            "性能安全风险",
            "供应商管理风险",
            "进度控制风险",
            "成本控制风险",
            "人力资源风险",
            "变更管理风险",
            "合规监管风险"
        ]
        
        # Create all sheets
        self.create_instruction_sheet()
        self.create_expert_scores_sheet(risk_factors)
        self.create_expert_weights_sheet()
        
        # Save template
        self.wb.save(output_path)


def main():
    """Generates the Excel template"""
    try:
        generator = ExcelTemplateGenerator()
        generator.generate_template("risk_evaluation.xlsx")
        print("Successfully generated risk evaluation template: risk_evaluation.xlsx")
        
    except Exception as e:
        print(f"Error generating template: {str(e)}")


if __name__ == "__main__":
    main()